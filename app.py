import os
import io
import uuid
import threading
import time
from flask import Flask, request, jsonify, render_template, send_file
import openpyxl

app = Flask(__name__)

MAX_CONTENT_LENGTH = 100 * 1024 * 1024  # 100MB tổng mỗi request
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH

ALLOWED_EXT = {".xlsx", ".xlsm", ".xltx", ".xltm"}

ERROR_VALUES = {
    "#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!",
    "#NULL!", "#SPILL!", "#CALC!", "#GETTING_DATA", "#FIELD!",
    "#BLOCKED!", "#UNKNOWN!"
}

# Lưu job scan trong bộ nhớ (đủ dùng cho team nhỏ dùng chung 1 instance Render).
# Mỗi job: {status, progress, results, message, files_meta}
JOBS = {}
JOBS_LOCK = threading.Lock()


def allowed_file(filename):
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXT


def detect_error(value, selected_errors):
    if value is None:
        return None
    text = str(value).strip().upper()
    if text in ERROR_VALUES and text in selected_errors:
        return text
    return None


def list_sheets_for_files(file_storages):
    """Đọc nhanh tên sheet của từng file để trả về cho frontend chọn."""
    files_info = []
    for f in file_storages:
        if not allowed_file(f.filename):
            continue
        content = f.read()
        f.seek(0)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            sheets = []
        files_info.append({
            "filename": f.filename,
            "sheets": sheets,
            "size": len(content)
        })
    return files_info


def run_scan(job_id, files_data, selected_sheets_map, selected_errors):
    """
    files_data: list of {"filename": str, "content": bytes}
    selected_sheets_map: { filename: [sheet_name, ...] }  (rỗng/không có key = quét tất cả sheet của file đó)
    selected_errors: set
    """
    with JOBS_LOCK:
        JOBS[job_id]["status"] = "running"

    try:
        results = []
        total_files = len(files_data)

        for file_index, fd in enumerate(files_data, start=1):
            filename = fd["filename"]
            content = fd["content"]

            with JOBS_LOCK:
                JOBS[job_id]["message"] = f"Đang mở: {filename}"

            try:
                wb_values = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)
                wb_formulas = openpyxl.load_workbook(io.BytesIO(content), data_only=False, read_only=False)
            except Exception as e:
                with JOBS_LOCK:
                    JOBS[job_id]["message"] = f"Lỗi mở file {filename}: {e}"
                continue

            wanted_sheets = selected_sheets_map.get(filename)
            if not wanted_sheets:
                wanted_sheets = wb_values.sheetnames

            for sheet_name in wanted_sheets:
                if sheet_name not in wb_values.sheetnames:
                    continue

                with JOBS_LOCK:
                    JOBS[job_id]["message"] = f"Đang quét: {filename} — {sheet_name}"

                ws_values = wb_values[sheet_name]
                ws_formulas = wb_formulas[sheet_name]

                max_row = ws_values.max_row
                max_col = ws_values.max_column
                if max_row == 0 or max_col == 0:
                    continue

                for row_values, row_formulas in zip(ws_values.iter_rows(), ws_formulas.iter_rows()):
                    for cell_value, cell_formula in zip(row_values, row_formulas):
                        value = cell_value.value
                        error = detect_error(value, selected_errors)
                        if error:
                            formula = ""
                            raw_formula = cell_formula.value
                            if isinstance(raw_formula, str) and raw_formula.startswith("="):
                                formula = raw_formula

                            result = {
                                "file": filename,
                                "sheet": sheet_name,
                                "row": cell_value.row,
                                "column": cell_value.column_letter,
                                "cell": cell_value.coordinate,
                                "error": error,
                                "formula": formula
                            }
                            results.append(result)

                            with JOBS_LOCK:
                                JOBS[job_id]["results"] = results
                                JOBS[job_id]["progress"] = int((file_index / total_files) * 100)

            with JOBS_LOCK:
                JOBS[job_id]["progress"] = int((file_index / total_files) * 100)

        with JOBS_LOCK:
            JOBS[job_id]["status"] = "finished"
            JOBS[job_id]["results"] = results
            JOBS[job_id]["progress"] = 100
            JOBS[job_id]["message"] = f"Hoàn tất. Tìm thấy {len(results)} lỗi."

    except Exception as e:
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["message"] = f"Scan thất bại: {e}"


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/list-sheets", methods=["POST"])
def api_list_sheets():
    """Nhận nhiều file, trả về danh sách sheet của từng file để user chọn."""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Không có file nào được gửi lên."}), 400

    valid_files = [f for f in files if allowed_file(f.filename)]
    if not valid_files:
        return jsonify({"error": "Không có file Excel hợp lệ (.xlsx, .xlsm, .xltx, .xltm)."}), 400

    files_info = list_sheets_for_files(valid_files)
    return jsonify({"files": files_info})


@app.route("/api/scan", methods=["POST"])
def api_scan():
    """
    Bắt đầu 1 job scan chạy nền. Nhận:
      - files: multipart files
      - selected_sheets: JSON string { filename: [sheet,...] }
      - selected_errors: JSON string [error,...]
    Trả về job_id để frontend poll tiến độ.
    """
    import json

    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Không có file nào được gửi lên."}), 400

    valid_files = [f for f in files if allowed_file(f.filename)]
    if not valid_files:
        return jsonify({"error": "Không có file Excel hợp lệ."}), 400

    try:
        selected_sheets_map = json.loads(request.form.get("selected_sheets", "{}"))
    except Exception:
        selected_sheets_map = {}

    try:
        selected_errors_list = json.loads(request.form.get("selected_errors", "[]"))
    except Exception:
        selected_errors_list = []

    selected_errors = set(e for e in selected_errors_list if e in ERROR_VALUES)
    if not selected_errors:
        return jsonify({"error": "Vui lòng chọn ít nhất 1 loại lỗi."}), 400

    files_data = [{"filename": f.filename, "content": f.read()} for f in valid_files]

    job_id = str(uuid.uuid4())
    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "queued",
            "progress": 0,
            "results": [],
            "message": "Đang khởi tạo...",
            "created_at": time.time()
        }

    thread = threading.Thread(
        target=run_scan,
        args=(job_id, files_data, selected_sheets_map, selected_errors),
        daemon=True
    )
    thread.start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def api_status(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({"error": "Job không tồn tại."}), 404
        return jsonify({
            "status": job["status"],
            "progress": job["progress"],
            "message": job["message"],
            "results": job["results"],
            "error_count": len(job["results"])
        })


@app.route("/api/export/<job_id>")
def api_export(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job or job["status"] != "finished":
            return jsonify({"error": "Job chưa hoàn tất hoặc không tồn tại."}), 400
        results = list(job["results"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Error Report"

    headers = ["File", "Sheet", "Row", "Column", "Cell", "Error", "Formula"]
    ws.append(headers)

    for r in results:
        ws.append([r["file"], r["sheet"], r["row"], r["column"], r["cell"], r["error"], r["formula"]])

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {"A": 28, "B": 20, "C": 8, "D": 10, "E": 10, "F": 16, "G": 70}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="Excel_Error_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Dọn job cũ (>2 giờ) để tránh phình bộ nhớ khi nhiều người dùng chung
def cleanup_old_jobs():
    while True:
        time.sleep(1800)
        cutoff = time.time() - 7200
        with JOBS_LOCK:
            stale = [jid for jid, j in JOBS.items() if j.get("created_at", 0) < cutoff]
            for jid in stale:
                del JOBS[jid]


threading.Thread(target=cleanup_old_jobs, daemon=True).start()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
